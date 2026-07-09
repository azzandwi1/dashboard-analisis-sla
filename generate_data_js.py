import openpyxl
import gzip
import base64
import io
import csv
import os

# CONFIGURATION
excel_path = "Percentile Achievement New.xlsx"
js_path = "data.js"

# If you add a new column/dimension to your PivotTables, update this count.
# Currently, there are 10 dimensions.
NUM_DIMENSIONS = 10 

def main():
    if not os.path.exists(excel_path):
        print(f"Error: File '{excel_path}' tidak ditemukan di folder ini.")
        return

    print(f"1. Membaca workbook '{excel_path}'...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if "Pivot_SLA" not in wb.sheetnames or "Pivot_Hari" not in wb.sheetnames:
        print("Error: Sheet 'Pivot_SLA' atau 'Pivot_Hari' tidak ditemukan di file Excel.")
        wb.close()
        return

    # ==========================================
    # STEP 1: READ PIVOT_SLA
    # ==========================================
    print("2. Membaca data SLA dari sheet 'Pivot_SLA'...")
    sheet_sla = wb["Pivot_SLA"]
    sla_dict = {}
    sla_iterator = sheet_sla.iter_rows(values_only=True)
    header_sla = next(sla_iterator)

    print("Header SLA yang dideteksi:", header_sla[:NUM_DIMENSIONS + 2])
    
    count_sla = 0
    for row in sla_iterator:
        if not row[0] or row[0] == "Grand Total":
            continue
        
        # Ambil key dimensi filter
        key = tuple(str(x).strip() if x is not None else "(blank)" for x in row[:NUM_DIMENSIONS])
        
        # Ambil nilai SLA Min dan Max (berada tepat setelah kolom dimensi)
        sla_min = row[NUM_DIMENSIONS] if row[NUM_DIMENSIONS] is not None else 0.0
        sla_max = row[NUM_DIMENSIONS + 1] if row[NUM_DIMENSIONS + 1] is not None else 0.0
        
        sla_dict[key] = (sla_min, sla_max)
        count_sla += 1

    print(f"   -> Berhasil memuat {count_sla} definisi rute SLA.")

    # ==========================================
    # STEP 2: READ PIVOT_HARI & MERGE
    # ==========================================
    print("3. Menggabungkan data dengan distribusi hari dari sheet 'Pivot_Hari'...")
    sheet_hari = wb["Pivot_Hari"]
    hari_iterator = sheet_hari.iter_rows(values_only=True)
    header_hari = next(hari_iterator)

    # Petakan kolom hari yang positif (>= 0)
    day_mapping = {}
    for idx in range(NUM_DIMENSIONS, len(header_hari)):
        col_val = header_hari[idx]
        if col_val is None:
            continue
        try:
            day_int = int(col_val)
            if day_int >= 0:
                day_mapping[idx] = day_int
        except ValueError:
            # Lewati kolom teks/kosong/blank
            continue

    print(f"   -> Mendeteksi {len(day_mapping)} kolom hari positif (H0, H1, dst).")

    # Siapkan CSV di memori
    csv_buffer = io.StringIO()
    csv_writer = csv.writer(csv_buffer)

    # Tulis Header CSV hasil gabungan
    csv_headers = list(header_sla[:NUM_DIMENSIONS]) + [
        "Weighted Avg SLA Min", "Weighted Avg SLA Max", "Total AWB", "Distribution"
    ]
    csv_writer.writerow(csv_headers)

    count_merged = 0
    count_skipped = 0

    for row in hari_iterator:
        if not row[0] or row[0] == "Grand Total":
            continue
            
        key = tuple(str(x).strip() if x is not None else "(blank)" for x in row[:NUM_DIMENSIONS])
        
        # Ambil SLA Min dan Max dari database Pivot_SLA
        sla_min, sla_max = sla_dict.get(key, (0.0, 0.0))
        
        # Hitung distribusi hari dan total AWB
        dist_parts = []
        total_awb = 0
        
        for idx, day_val in day_mapping.items():
            if idx < len(row):
                val = row[idx]
                if val is not None and val > 0:
                    dist_parts.append(f"{day_val}:{val}")
                    total_awb += val
                    
        # Lewati baris jika total paket AWB adalah 0
        if total_awb == 0:
            count_skipped += 1
            continue
            
        dist_str = ";".join(dist_parts)
        
        # Gabungkan data ke CSV
        csv_row = list(key) + [sla_min, sla_max, total_awb, dist_str]
        csv_writer.writerow(csv_row)
        count_merged += 1

    # ==========================================
    # STEP 3: COMPRESS & WRITE TO DATA.JS
    # ==========================================
    print("4. Mengompresi data dengan GZIP dan mengodekan ke Base64...")
    csv_text = csv_buffer.getvalue()
    csv_bytes = csv_text.encode('utf-8')
    
    compressed_bytes = gzip.compress(csv_bytes)
    base64_str = base64.b64encode(compressed_bytes).decode('utf-8')

    print(f"5. Menulis hasil kompresi ke file '{js_path}'...")
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write("const COMPRESSED_DATA = `")
        f.write(base64_str)
        f.write("`;\n")

    print("\nPROSES SELESAI DENGAN SUKSES!")
    print(f"   - Total baris yang digabungkan: {count_merged}")
    print(f"   - Baris 0 AWB yang diabaikan: {count_skipped}")
    print(f"   - Ukuran CSV Asli: {len(csv_bytes) / (1024*1024):.2f} MB")
    print(f"   - Ukuran File data.js Baru: {len(base64_str) / (1024*1024):.2f} MB")
    
    wb.close()

if __name__ == "__main__":
    main()
